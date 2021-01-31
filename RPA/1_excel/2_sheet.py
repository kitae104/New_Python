from openpyxl import Workbook
wb = Workbook()             # 새 워크북 생성
ws = wb.create_sheet()      # 새로운 sheet 기본 이름으로 생성
ws.title = "MySheet"        # 시트 이름 변경
ws.sheet_properties.tabColor = "ff3399"

# Sheet, MySheet, YourSheet
ws1 = wb.create_sheet("YourSheet")  # 주어진 이름으로 Sheet 생성
ws2 = wb.create_sheet("NewSheet", 2)  # 주어진 이름으로 Sheet 생성

new_ws = wb["NewSheet"]     # Dict 형태로 시트에 접근

print(wb.sheetnames)

# 시트 복사
new_ws["A1"] = "Test"               # 특정 위치에 데이터 입력
target = wb.copy_worksheet(new_ws)  # 워크시트 복사하기
target.title = "Copied Sheet"       # 타이틀 변경

wb.save("sample.xlsx")
wb.close()