# 셀 다루기
from openpyxl import Workbook
wb = Workbook()             # 새 워크북 생성
ws = wb.active
ws.title = "기태_자료"

ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3
ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"].value)

wb.save("sample.xlsx")
wb.close()