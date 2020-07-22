# -*- coding: utf-8 -*-

import sys
import PyQt5
from PyQt5.QtGui import *
from PyQt5.QtCore import *
from PyQt5.QtWidgets import *
from PyQt5 import uic
import xlrd
import openpyxl
import time
from selenium import webdriver
from Activity import *
import keyboard

# Qt디자이너로 만든 UI파일
Gui = r'.\designer\GUI.ui'


# 메인 윈도우창
class MainWindow(QMainWindow):

    # 초기화
    def __init__(self):
        QDialog.__init__(self, None)
        uic.loadUi(Gui, self)  # UI 불러오기
        self.activity = Activity()  # Activity.py 객체 생성
        self.activity_list = []  # 프로퍼티 리스트 초기화
        self.row_num = 0  # 현재 줄번호 변수 초기화
        self.ActivityBox.expandAll()  # 트리위젯 펼치기
        self.my_item = -1   # 사용자함수 갯수 변수 초기화
        self.my_num = 0  # 현재 사용자함수 변수 초기화
        self.my_text = []  # 사용자함수 리스트 초기화

    # 메뉴얼열기
    def LogoClick(self):
        os.startfile('메뉴얼.pdf')  # 메뉴얼 열기

    # 새로만들기
    def NewFile(self):
        uic.loadUi(Gui, self)  # UI 불러오면서 초기화
        self.activity_list = []  # 프로퍼티 리스트 초기화
        self.row_num = 0  # 현재 줄번호 변수
        self.ActivityBox.expandAll()  # 트리위젯 펼치기
        self.my_item = -1  # 사용자함수 갯수 변수 초기화
        self.my_num = 0  # 현재 사용자함수 변수 초기화
        self.my_text = []  # 사용자함수 리스트 초기화

    # 파일 열기
    def OpenFile(self):
        name = QFileDialog.getOpenFileName()  # 파일열기 다이얼로그 실행
        if name[0] != '':  # 취소를 안눌렀을 경우
            self.MainBox.clearContents()  # 메인박스 초기화
            self.activity_list = []  # 프로퍼티 초기화
            self.row_num = 0  # 현재 줄번호 변수 초기화
            self.my_item = -1  # 사용자함수 갯수 변수 초기화
            self.my_num = 0  # 현재 사용자함수 변수 초기화
            self.my_text = []  # 사용자함수 리스트 초기화

            wb = xlrd.open_workbook(name[0])  # 엑셀파일 열기
            ws = wb.sheet_by_index(0)  # 첫번째 시트 메인박스
            ws2 = wb.sheet_by_index(1)  # 두번째 시트 프로퍼티
            ws3 = wb.sheet_by_index(2)  # 세번째 시트 사용자함수

            numrow = ws.nrows  # 첫번째 시트 row
            numcol = ws.ncols  # 첫번째 시트 col
            num2row = ws2.nrows  # 두번째 시트 row
            num2col = ws2.ncols  # 두번째 시트 col
            num3row = ws3.nrows  # 세번째 시트 row
            num3col = ws3.ncols  # 세번째 시트 col

            self.MainBox.setColumnCount(numcol)  # 메인박스 col 생성
            self.MainBox.setRowCount(numrow)  # 메인박스 row 생성

            # 메인 박스에 데이터 입력
            for row in range(numrow):
                for col in range(numcol):
                    if ws.row_values(row)[col]:  # 첫번째 시트 해당칸에 값이 있을경우
                        self.MainBox.setItem(row, col, QTableWidgetItem(ws.row_values(row)[col]))  # 메인박스 같은칸에 추가

            # 프로퍼티 리스트에 추가
            for row2 in range(num2row):
                for col2 in range(num2col):
                    if ws2.row_values(row2)[col2]:  # 두번째 시트 해당칸에 값이 있을경우
                        self.activity_list.append(ws2.row_values(row2)[col2])  # 프로퍼티 리스트에 추가

            # 리스트에 추가된 번호 원상태로 변경
            for i in range(numrow):
                self.activity_list[self.activity_list.index(i + 1)] = i  # 저장할 때 1씩 늘려놓았던 줄번호 복구
            self.ShowProperty(self.row_num)  # 첫번째 줄 기능의 프로퍼티 보이기

            # 사용자함수 리스트에 추가
            for row3 in range(num3row):
                for col3 in range(num3col):
                    if ws3.row_values(row3)[col3]:  # 세번째 시트 해당칸에 값이 있을경우
                        self.my_text.append(ws3.row_values(row3)[col3])  # 사용자함수 리스트에 추가
            self.my_item = len(self.my_text)  # 사용자함수 갯수 변수 값 변경

    # 파일 저장
    def SaveFile(self):
        if self.MainBox.model().rowCount():  # 메인박스에 데이터가 있을 경우
            name = QFileDialog.getSaveFileName()  # 파일저장 다이얼로그 실행
            if name[0] != '':  # 취소를 안눌렀을 경우
                self.SetProperty(self.row_num)  # 현재 선택된 기능의 프로퍼티 저장
                self.SaveMyItem()  # 사용자함수 저장
                numrow = self.MainBox.model().rowCount()  # 메인박스의 row 갯수
                numcol = self.MainBox.model().columnCount()  # 메인박스의 col 갯수

                wb = openpyxl.Workbook()  # 워크북 생성
                ws = wb.active  # 첫번째 워크시트 생성 메인박스
                # 메인박스 아이템들 저장
                for row in range(numrow):
                    for col in range(numcol):
                        if self.MainBox.item(row, col) is not None:  # 메인박스에 값이 비어있지 않을 때
                            ws.cell(row=row + 1, column=col + 1).value = self.MainBox.item(row, col).text()  # 첫번째 워크시트 해당칸에 저장
                        else:  # 메인박스에 값이 비어있을 때
                            ws.cell(row=row + 1, column=col + 1).value = ''  # 첫번째 워크시트 해당칸에 공백저장

                sheet2 = wb.create_sheet("var")  # 두번째 워크시트 생성 프로퍼티

                for i in range(numrow, 0, -1):  # 줄번호 역순으로 내려가기
                    self.activity_list[self.activity_list.index(i-1)] = i  # 프로퍼티 리스트에 있는 줄번호들 +1씩하기(0은 데이터로 입력이안됨)
                for i in range(self.activity_list.count('')):  # 프로퍼티가 빈칸일 때
                    self.activity_list[self.activity_list.index('')] = '_'  # 빈칸을 _로 채우기(빈칸은 데이터로 입력이안됨)
                sheet2.append(self.activity_list)  # 두번째 시트에 프로퍼티 리스트 저장

                # 사용자함수 저장하기
                sheet3 = wb.create_sheet("myitem")  # 세번째 워크시트 생성 사용자 함수
                sheet3.append(self.my_text)  # 세번째 시트에 사용자함수 저장

                wb.save(name[0] + '.xlsx')  # 파일 저장하기에서 정한 경로와 이름에 .xlsx 확장자를 붙여서 저장

                for i in range(numrow):
                    self.activity_list[self.activity_list.index(i + 1)] = i  # 저장 후 프로퍼티 리스트의 줄번호를 복구

    # 실행
    def Run(self):  # 동작 함수
        if self.MainBox.model().rowCount():  # 메인박스에 데이터가 있을 경우
            self.SetProperty(self.row_num)  # 현재 선택된 기능의 프로퍼티 저장
            self.SaveMyItem()  # 사용자 함수 저장
            for row in range(self.MainBox.model().rowCount()):  # 한줄씩 내려가면서 기능 실행
                if self.Action(row) == 0:  # 141번째 줄 동작 과정함수에서 오류가 났을 경우
                    break  # 동작 중지

    def Action(self, row):  # 동작 과정함수
        try:  # 에러가 안날 경우
            if self.MainBox.item(row, 0).text() == '브라우저 열기':  # 현재 줄의 기능의 이름이 브라우저 열기일 경우 (다른기능 전부 같음)
                Activity.OpenBrowser(self, self.activity_list[self.activity_list.index(row) + 2])  # Activity.py에 있는 기능 실행 (프로퍼티 변수는 리스트에 줄번호, 변수이름, 변수 값의 순으로 저장되므로 줄번호 + 2)
            elif self.MainBox.item(row, 0).text() == 'id값으로 클릭':
                Activity.find_element_by_id_click(self, self.activity_list[self.activity_list.index(row) + 2])
            elif self.MainBox.item(row, 0).text() == 'id값으로 입력':
                Activity.find_element_by_id_sendkey(self, self.activity_list[self.activity_list.index(row) + 2],
                                                    self.activity_list[self.activity_list.index(row) + 4])
            elif self.MainBox.item(row, 0).text() == 'class값으로 클릭':
                Activity.find_element_by_class_click(self, self.activity_list[self.activity_list.index(row) + 2])
            elif self.MainBox.item(row, 0).text() == 'class값으로 입력':
                Activity.find_element_by_class_sendkey(self, self.activity_list[self.activity_list.index(row) + 2],
                                                       self.activity_list[self.activity_list.index(row) + 4])
            elif self.MainBox.item(row, 0).text() == 'css값으로 클릭':
                Activity.find_element_by_css_click(self, self.activity_list[self.activity_list.index(row) + 2])
            elif self.MainBox.item(row, 0).text() == 'css값으로 입력':
                Activity.find_element_by_css_sendkey(self, self.activity_list[self.activity_list.index(row) + 2],
                                                     self.activity_list[self.activity_list.index(row) + 4])
            elif self.MainBox.item(row, 0).text() == 'xpath값으로 클릭':
                Activity.find_element_by_xpath_click(self, self.activity_list[self.activity_list.index(row) + 2])
            elif self.MainBox.item(row, 0).text() == 'xpath값으로 입력':
                Activity.find_element_by_xpath_sendkey(self, self.activity_list[self.activity_list.index(row) + 2],
                                                       self.activity_list[self.activity_list.index(row) + 4])
            elif self.MainBox.item(row, 0).text() == 'Gmail 세션 생성':
                Activity.Gmail_Session_Create(self, self.activity_list[self.activity_list.index(row) + 2],
                                              self.activity_list[self.activity_list.index(row) + 4])
            elif self.MainBox.item(row, 0).text() == '파일첨부':
                Activity.Attach_File(self, self.activity_list[self.activity_list.index(row) + 2])
            elif self.MainBox.item(row, 0).text() == '엑셀 읽어 보내기':
                Activity.Excel_read_send(self, self.activity_list[self.activity_list.index(row) + 2])
            elif self.MainBox.item(row, 0).text() == 'text 크롤링':
                Activity.WebCrawling_text(self, self.activity_list[self.activity_list.index(row) + 2])
            elif self.MainBox.item(row, 0).text() == 'url 크롤링':
                Activity.WebCrawling_url(self, self.activity_list[self.activity_list.index(row) + 2],
                                         self.activity_list[self.activity_list.index(row) + 4])
            elif self.MainBox.item(row, 0).text() == '데이터로 크롤링':
                Activity.WebCrawling_get_url(self)
            elif self.MainBox.item(row, 0).text() == 'url 여러 페이지 크롤링':
                Activity.WebCrawling_url_count(self, self.activity_list[self.activity_list.index(row) + 2],
                                               self.activity_list[self.activity_list.index(row) + 4],
                                               self.activity_list[self.activity_list.index(row) + 6])
            elif self.MainBox.item(row, 0).text() == '텍스트 파일 읽기':
                Activity.Read_TextFile(self, self.activity_list[self.activity_list.index(row) + 2])
            elif self.MainBox.item(row, 0).text() == '텍스트 파일 쓰기':
                Activity.Write_TextFile(self, self.activity_list[self.activity_list.index(row) + 2])
            elif self.MainBox.item(row, 0).text() == '텍스트 추가':
                Activity.Add_Text(self, self.activity_list[self.activity_list.index(row) + 2],
                                  self.activity_list[self.activity_list.index(row) + 4])
            elif self.MainBox.item(row, 0).text() == '텍스트 파일을 엑셀 파일로 변환':
                Activity.txt_to_xlsx(self, self.activity_list[self.activity_list.index(row) + 2])
            elif self.MainBox.item(row, 0).text() == '시트 불러오기':
                Activity.Load_Sheet(self, self.activity_list[self.activity_list.index(row) + 2])
            elif self.MainBox.item(row, 0).text() == '행 삭제':
                Activity.Delete_rows(self, self.activity_list[self.activity_list.index(row) + 2])
            elif self.MainBox.item(row, 0).text() == '열 삭제':
                Activity.Delete_cols(self, self.activity_list[self.activity_list.index(row) + 2])
            elif self.MainBox.item(row, 0).text() == '셀 옮기기':
                Activity.Move_Cell(self, self.activity_list[self.activity_list.index(row) + 2],
                                   self.activity_list[self.activity_list.index(row) + 4])
            elif self.MainBox.item(row, 0).text() == '값 얻기':
                Activity.GetValue(self, self.activity_list[self.activity_list.index(row) + 2])
            elif self.MainBox.item(row, 0).text() == '행 찾기':
                Activity.Find_row(self, self.activity_list[self.activity_list.index(row) + 2],
                                  self.activity_list[self.activity_list.index(row) + 4],
                                  self.activity_list[self.activity_list.index(row) + 6])
            elif self.MainBox.item(row, 0).text() == '값 설정':
                Activity.SetValue(self, self.activity_list[self.activity_list.index(row) + 2],
                                  self.activity_list[self.activity_list.index(row) + 4])
            elif self.MainBox.item(row, 0).text() == 'docx 읽기':
                Activity.Read_docx(self, self.activity_list[self.activity_list.index(row) + 2])
            elif self.MainBox.item(row, 0).text() == 'docx 생성':
                Activity.Make_docx(self, self.activity_list[self.activity_list.index(row) + 2])
            elif self.MainBox.item(row, 0).text() == '이미지 추가':
                Activity.Add_Image(self, self.activity_list[self.activity_list.index(row) + 2],
                                   self.activity_list[self.activity_list.index(row) + 4])
            elif self.MainBox.item(row, 0).text() == 'docx 작성':
                Activity.Make_docx(self, self.activity_list[self.activity_list.index(row) + 2])
            elif self.MainBox.item(row, 0).text() == 'head 쓰기':
                Activity.Write_head(self, self.activity_list[self.activity_list.index(row) + 2],
                                    self.activity_list[self.activity_list.index(row) + 4])
            elif self.MainBox.item(row, 0).text() == 'body 쓰기':
                Activity.Write_body(self, self.activity_list[self.activity_list.index(row) + 2],
                                    self.activity_list[self.activity_list.index(row) + 4])
            elif self.MainBox.item(row, 0).text() == 'docx 저장':
                Activity.Save_docx(self)
            elif self.MainBox.item(row, 0).text() == 'csv 읽기':
                Activity.Read_csv(self, self.activity_list[self.activity_list.index(row) + 2])
            elif self.MainBox.item(row, 0).text() == '열 가져오기':
                Activity.Get_Line(self, self.activity_list[self.activity_list.index(row) + 2])
            elif self.MainBox.item(row, 0).text() == 'csv 쓰기':
                Activity.Write_csv(self, self.activity_list[self.activity_list.index(row) + 2])
            elif self.MainBox.item(row, 0).text() == '행 추가':
                Activity.Add_row(self, self.activity_list[self.activity_list.index(row) + 2])
            elif self.MainBox.item(row, 0).text() == 'csv파일 추가모드로 열기':
                Activity.Open_Add_mode(self, self.activity_list[self.activity_list.index(row) + 2])
            elif self.MainBox.item(row, 0).text() == '파일 닫기':
                Activity.File_Close(self)
            elif self.MainBox.item(row, 0).text() == '모든 값 얻기':
                Activity.Get_All_Value(self)
            elif self.MainBox.item(row, 0).text() == '팝업창 닫기':
                Activity.Close_PopUp(self)
            elif self.MainBox.item(row, 0).text() == '브라우저 닫기':
                Activity.Close_Browser(self)
            elif self.MainBox.item(row, 0).text() == '이미지 다운로드(src)':
                Activity.Image_Download(self, self.activity_list[self.activity_list.index(row) + 2])
            elif self.MainBox.item(row, 0).text() == '페이지 스크린샷':
                Activity.Page_Screenshot(self)
            elif self.MainBox.item(row, 0).text() == '이미지 클릭':
                Activity.Image_Click(self, self.activity_list[self.activity_list.index(row) + 2])
            elif self.MainBox.item(row, 0).text() == '좌표 클릭':
                Activity.Position_Click(self, self.activity_list[self.activity_list.index(row) + 2],
                                        self.activity_list[self.activity_list.index(row) + 4])
            elif self.MainBox.item(row, 0).text() == '딜레이':
                Activity.Delay(self, self.activity_list[self.activity_list.index(row) + 2])
            elif self.MainBox.item(row, 0).text() == '반복_For':
                self.For_(self.activity_list[self.activity_list.index(row) + 2],
                          self.activity_list[self.activity_list.index(row) + 4],
                          self.activity_list[self.activity_list.index(row) + 6])
            elif 'MyItem' in self.MainBox.item(row, 0).text():  # 사용자함수 실행
                self.RunMyItem(row)  # 663번째 줄 함수 실행
        except:  # 에러가 났을 경우
            QMessageBox.warning(self, 'ERROR', self.MainBox.item(row, 0).text()+' 에러\n'+'다시 확인해주세요.')  # 에러메세지 표시
            return 0  # 0을 리턴

    # 종료
    def Stop(self):
        sys.exit()

    # 한칸 위로 올리기
    def Up(self):
        try:  # 에러가 안날 경우
            row = self.MainBox.currentRow()  # 메인박스의 현재 row
            column = self.MainBox.currentColumn()  # 메인박스의 현재 col
            if row == -1:  # row값이 -1일 때(삭제과정을 거치면 현재 row가 -1이 됨)
                pass
            elif row > 0:  # 현재 row가 있을 때
                # 메인박스
                self.MainBox.insertRow(row - 1)  # 한줄 위에 줄 삽입
                for i in range(self.MainBox.columnCount()):  # col 갯수만큼(2번)
                    self.MainBox.setItem(row - 1, i, self.MainBox.takeItem(row + 1, i))  # 옮길 아이템 선택후 삽입
                    self.MainBox.setCurrentCell(row - 1, column)  # 현재 셀의 값을 한줄 위로 변경
                self.MainBox.removeRow(row + 1)  # 기존에 있었던 줄을 삭제
                # 프로퍼티
                now = self.activity_list.index(row-1)  # 프로퍼티 리스트에서 옮긴 후의 줄번호
                if row + 1 not in self.activity_list:  # 기존의 값이 프로퍼티 리스트에서 마지막 번호였을 때
                    for i in range(self.activity_list.index(row), len(self.activity_list)):  # 리스트에서 row의 위치부터 끝까지
                        self.activity_list.insert(now, self.activity_list[i])  # now의 위치에 기존의 값을 삽입
                        del(self.activity_list[i+1])  # 기존에 있던 값을 삭제
                        now = now + 1  # now 위치 변경을 위해 +1
                else:
                    for i in range(self.activity_list.index(row), self.activity_list.index(row+1)):  #리스트에서 row의 위치부터 row+1 까지
                        self.activity_list.insert(now, self.activity_list[i])  # now의 위치에 기존의 값을 삽입
                        del (self.activity_list[i + 1])  # 기존에 있던 값을 삭제
                        now = now + 1  # now 위치 변경을 위해 +1
                self.activity_list[self.activity_list.index(row - 1)] = row  # 옮겨진 위치에 있던 줄번호 값을 변경
                self.activity_list[self.activity_list.index(row)] = row - 1  # 옮긴 줄번호 값을 변경
                self.row_num = self.row_num - 1  # 현재 줄번호 변경
        except:
            QMessageBox.warning(self, 'ERROR', 'Up 에러')  # 에러메세지 표시

    # 한칸 아래로 내리기
    def Down(self):
        try:
            row = self.MainBox.currentRow()  # 현재 row의 값
            column = self.MainBox.currentColumn()  # 현재 col의 값
            if row == -1:  # row값이 -1일 때(삭제과정을 거치면 현재 row가 -1이 됨)
                pass
            elif row < self.MainBox.rowCount() - 1:  # row 밑에 다른 값들이 있을 때
                # 메인박스
                self.MainBox.insertRow(row + 2)  # 옮길 위치에 줄 삽입
                for i in range(self.MainBox.columnCount()):  # col 갯수만큼(2번)
                    self.MainBox.setItem(row + 2, i, self.MainBox.takeItem(row, i))  # 옮길 아이템 선택후 삽입
                    self.MainBox.setCurrentCell(row + 2, column)  # 현재 셀의 값을 아래로 변경
                self.MainBox.removeRow(row)  # 기존에 있던 줄 삭제
                # 프로퍼티
                if row + 2 not in self.activity_list:  # 프로퍼티 리스트에 다음줄이 마지막일 경우
                    for i in range(self.activity_list.index(row), self.activity_list.index(row+1)):  # 현재 프로퍼티 리스트에서 줄번호와 변수이름, 값만큼
                        self.activity_list.append(self.activity_list[i])  # 맨뒤에 삽입하기
                else:  # 다음줄이 마지막이 아닐경우
                    now = self.activity_list.index(row+2)  # 옮겨질 위치
                    for i in range(self.activity_list.index(row), self.activity_list.index(row+1)):  # 현재 프로퍼티 리스트에서 줄번호와 변수이름, 값만큼
                        self.activity_list.insert(now, self.activity_list[i])  # 옮겨질 위치에 삽입하기
                        now = now + 1  # 옮겨질 위치 변경
                del (self.activity_list[self.activity_list.index(row):self.activity_list.index(row + 1)])  # 기존에 있던 값들 삭제
                self.activity_list[self.activity_list.index(row)] = row + 1  # 옮긴 줄번호 값을 변경
                self.activity_list[self.activity_list.index(row + 1)] = row  # 옮겨진 위치에 있던 줄번호 값을 변경
                self.row_num = self.row_num + 1  # 현재 줄번호 변경
        except:
            QMessageBox.warning(self, 'ERROR', 'Down 에러')  # 에러메세지 표시

    # 삭제하기
    def Delete(self):
        if self.MainBox.currentItem() is None:  # 메인박스에 아이템이 없을 때
            pass
        elif self.MainBox.rowCount() == 1:  # 메인박스에 아이템이 한개 밖에 없을 때
            self.MainBox.setRowCount(0)  # 메인박스 줄 갯수 0으로 만들기
            self.varBox.setRowCount(0)  # 프로퍼티 박스 줄 갯수 0으로 만들기
            self.activity_list = []  # 프로퍼티 리스트 초기화
            self.row_num = 0  # 현재 줄번호 초기화
            self.my_item = -1  # 사용자함수 갯수 변수 초기화
            self.my_num = 0  # 현재 사용자함수 변수 초기화
            self.my_text = []  # 사용자함수 리스트 초기화
        else:  # 두개이상의 아이템이 있을 때
            self.MainBox.removeRow(self.MainBox.currentRow())  # 현재 줄 삭제
            if self.row_num + 1 not in self.activity_list:  # 다음 줄이 없을 때
                del (self.activity_list[self.activity_list.index(self.row_num):])  # 프로퍼티 리스트에서 현재 줄번호부터 마지막까지 삭제
            else:  # 다음 줄이 있을 때
                del (self.activity_list[self.activity_list.index(self.row_num):self.activity_list.index(self.row_num + 1)])  # 프로퍼티 리스트에서 현재 줄번호부터 다음 줄번호 전까지 삭제
                for i in range(self.row_num+1, self.MainBox.rowCount()+1):  # 다음줄번호 부터 마지막 줄번호까지
                    self.activity_list[self.activity_list.index(i)] = i - 1  # 각각 줄번호 -1
            if self.row_num == 0:  # 삭제한 줄이 첫번째 줄일 때
                pass
            else:  # 첫번째 줄이 아닐 때
                self.row_num = self.row_num - 1  # 현재줄을 윗줄로 변경
            self.ShowProperty(self.row_num)  # 윗줄의 프로퍼티 보여주기

    # 트리위젯에서 더블클릭으로 메인박스에 추가하기
    def AddItem(self):
        if self.ActivityBox.currentItem().text(1) != '':  # 변수칸이 비었을 때(목록이 추가되는걸 막기위해 추가)
            if self.MainBox.rowCount() > 0:  # 메인박스에 데이터가 있을 때
                self.MainBox.insertRow(self.MainBox.rowCount())  # 제일 밑에 한줄 추가
            else:  # 메인박스에 데이터가 없을 때
                self.MainBox.setRowCount(1)  # 메인박스의 줄 갯수를 1로 변경
            self.MainBox.setItem(self.MainBox.rowCount() - 1, 0, QTableWidgetItem(self.ActivityBox.currentItem().text(0)))  # 트리위젯에서 선택한 아이템의 이름을 가져와서 삽입
            self.MainBox.setItem(self.MainBox.rowCount() - 1, 1, QTableWidgetItem(self.ActivityBox.currentItem().text(1)))  # 트리위젯에서 선택한 아이템의 변수를 가져와서 삽입
            self.AppendTable(self.ActivityBox.currentItem().text(0))  # 381번째 줄 프로퍼티 리스트에 추가하는 함수 실행
            self.MainBox.setCurrentCell(self.MainBox.rowCount() - 1, 0)  # 추가한 아이템으로 현재 선택된 셀 값을 변경
            if self.varBox.rowCount() > 0:  # 프로퍼티 박스에 줄 갯수가 0이 아닐때
                self.SetProperty(self.row_num)  # 기존에 있던 프로퍼티 값을 저장
            if 'MyItem' in self.MainBox.item(self.row_num, 0).text():  # 기존에 선택된 아이템이 사용자 함수일 때
                self.SaveMyItem()  # 사용자함수 저장
            self.plainTextEdit.clear()  # 사용자함수 정의 칸 초기화
            self.row_num = self.MainBox.currentRow()  # 줄번호 변수를 현재 줄의 값으로 변경
            self.selectPosition()  # 609번째 줄 마우스 위치 지정 기능을 추가할 때 실행
            self.seleteFile()  # 537번째 줄 프로퍼티에 파일이 들어갈때 파일열기 다이얼로그 실행
            self.ViewExcelImage()  # 545번째 줄 프로퍼티의 값이 엑셀이거나 사진파일일때 미리보기기능 실행
            self.ShowProperty(self.MainBox.rowCount()-1)  # 추가한 기능의 프로퍼티 보여주기

    # 프로퍼티를 리스트로 구현해서 넣기
    def AppendTable(self, activity_name):
        self.activity_list.append(self.MainBox.rowCount()-1)  # 줄번호 삽입
        if activity_name == '브라우저 열기':  # 기능의 이름이 브라우저 열기 일때
            self.AppendList(activity_name)  # 483번째 줄의 기능 실행
        elif activity_name == 'id값으로 클릭':
            self.AppendList(activity_name)
        elif activity_name == 'id값으로 입력':
            self.AppendList(activity_name)
        elif activity_name == 'class값으로 클릭':
            self.AppendList(activity_name)
        elif activity_name == 'class값으로 입력':
            self.AppendList(activity_name)
        elif activity_name == 'css값으로 클릭':
            self.AppendList(activity_name)
        elif activity_name == 'css값으로 입력':
            self.AppendList(activity_name)
        elif activity_name == 'xpath값으로 클릭':
            self.AppendList(activity_name)
        elif activity_name == 'xpath값으로 입력':
            self.AppendList(activity_name)
        elif activity_name == 'Gmail 세션 생성':
            self.AppendList(activity_name)
        elif activity_name == '파일첨부':
            self.AppendList(activity_name)
        elif activity_name == '엑셀 읽어 보내기':
            self.AppendList(activity_name)
        elif activity_name == 'text 크롤링':
            self.AppendList(activity_name)
        elif activity_name == 'url 크롤링':
            self.AppendList(activity_name)
        elif activity_name == '데이터로 크롤링':
            self.AppendList(activity_name)
        elif activity_name == 'url 여러 페이지 크롤링':
            self.AppendList(activity_name)
        elif activity_name == '텍스트 파일 읽기':
            self.AppendList(activity_name)
        elif activity_name == '텍스트 파일 쓰기':
            self.AppendList(activity_name)
        elif activity_name == '텍스트 추가':
            self.AppendList(activity_name)
        elif activity_name == '텍스트 파일을 엑셀 파일로 변환':
            self.AppendList(activity_name)
        elif activity_name == '시트 불러오기':
            self.AppendList(activity_name)
        elif activity_name == '행 삭제':
            self.AppendList(activity_name)
        elif activity_name == '열 삭제':
            self.AppendList(activity_name)
        elif activity_name == '셀 옮기기':
            self.AppendList(activity_name)
        elif activity_name == '값 얻기':
            self.AppendList(activity_name)
        elif activity_name == '행 찾기':
            self.AppendList(activity_name)
        elif activity_name == '값 설정':
            self.AppendList(activity_name)
        elif activity_name == 'docx 읽기':
            self.AppendList(activity_name)
        elif activity_name == 'docx 생성':
            self.AppendList(activity_name)
        elif activity_name == '이미지 추가':
            self.AppendList(activity_name)
        elif activity_name == 'docx 작성':
            self.AppendList(activity_name)
        elif activity_name == 'head 쓰기':
            self.AppendList(activity_name)
        elif activity_name == 'body 쓰기':
            self.AppendList(activity_name)
        elif activity_name == 'docx 저장':
            self.AppendList(activity_name)
        elif activity_name == 'csv 읽기':
            self.AppendList(activity_name)
        elif activity_name == '열 가져오기':
            self.AppendList(activity_name)
        elif activity_name == 'csv 쓰기':
            self.AppendList(activity_name)
        elif activity_name == '행 추가':
            self.AppendList(activity_name)
        elif activity_name == 'csv파일 추가모드로 열기':
            self.AppendList(activity_name)
        elif activity_name == '파일 닫기':
            self.AppendList(activity_name)
        elif activity_name == '모든 값 얻기':
            self.AppendList(activity_name)
        elif activity_name == '팝업창 닫기':
            self.AppendList(activity_name)
        elif activity_name == '브라우저 닫기':
            self.AppendList(activity_name)
        elif activity_name == '이미지 다운로드(src)':
            self.AppendList(activity_name)
        elif activity_name == '페이지 스크린샷':
            self.AppendList(activity_name)
        elif activity_name == '이미지 클릭':
            self.AppendList(activity_name)
        elif activity_name == '좌표 클릭':
            self.AppendList(activity_name)
        elif activity_name == '딜레이':
            self.AppendList(activity_name)
        elif activity_name == '반복_For':
            self.AppendList(activity_name)

    # 리스트에 넣는 동작
    def AppendList(self, name):
        for i in range(self.activity.count_keys(name)):  # Activity.py의 197번째 줄 0부터 키의 갯수만큼 반복
            key = self.activity.search_name_key(name, i)  # Activity.py의 201번째 줄 이름으로 키 찾기
            self.activity_list.append(key)  # 프로퍼티 리스트에 key를 추가
            self.activity_list.append(self.activity.search_key_value(name, key))  # 프로퍼티 리스트에 해당 key의 값을 추가

    # 프로퍼티 보여주기
    def ShowProperty(self, row_num):
        j = 0
        if row_num + 1 not in self.activity_list:  # 해당 줄이 마지막 줄일 때
            new_num = len(self.activity_list)-1  # 프로퍼티 리스트의 전체길이 -1
        else:
            new_num = self.activity_list.index(row_num + 1)-1  # 다음줄값의 직전 위치
        if self.activity_list.index(row_num) == new_num:  # row_num과 new_num이 같을 때
            self.varBox.setRowCount(0)  # 프로퍼티값이 없음
        for i in range(self.activity_list.index(row_num), new_num):  # 프로퍼티 리스트에서 현재줄번호의 위치부터 new_num까지
            self.varBox.setRowCount(j // 2 + 1)  # 갯수에 맞게 줄 생성
            self.varBox.setItem(j // 2, j % 2, QTableWidgetItem(self.activity_list[i+1]))  # 해당 위치에 프로퍼티리스트의 값을 삽입
            j = j + 1

    # 수정된 프로퍼티 리스트에 저장하기
    def SetProperty(self, row_num):
        k = 0
        if row_num + 1 not in self.activity_list:  # 해당 줄이 마지막 줄일 때
            num = len(self.activity_list)  # 프로퍼티 리스트의 전체길이
        else:
            num = self.activity_list.index(self.row_num + 1)  # 프로퍼티 리스트에서 다음 줄번호의 위치
        for i in range(self.activity_list.index(self.row_num)+1, num):  # 현재줄번호 다음위치부터 다음 줄번호 까지
            self.activity_list[i] = self.varBox.item(k // 2, k % 2).text()  # 프로퍼티 리스트에 값 삽입
            k = k + 1

    # 클릭으로 프로퍼티 보이게하기
    def OnProperty(self):
        self.SetProperty(self.row_num)  # 기존에 선택됐던 기능의 프로퍼티값 저장
        if 'MyItem' in self.MainBox.item(self.row_num, 0).text():  # 기존에 선택됐던 기능이 사용자함수일 때
            self.SaveMyItem()  # 사용자함수 저장
        self.plainTextEdit.clear()  # 사용자함수 정의 칸 초기화
        self.row_num = self.MainBox.currentRow()  # 줄번호 변수를 현재 줄번호로 변경
        self.seleteFile()  # 537번째 줄 프로퍼티에 파일이 들어갈때 파일열기 다이얼로그 실행
        self.ViewExcelImage()  # 545번째 줄 프로퍼티의 값이 엑셀이거나 사진파일일때 미리보기기능 실행
        self.ShowProperty(self.row_num)  # 현재 기능의 프로퍼티 보여주기
        if 'MyItem' in self.MainBox.item(self.row_num, 0).text():  # 현재 선택한 기능이 사용자함수일 때
            string = self.MainBox.item(self.row_num, 0).text()  # 변수에 기능이름 저장
            self.my_num = int(string[6:])  # 몇번째 사용자 함수인지 my_num에 저장
            self.ShowMyItem()  # 해당 기능의 사용자 정의 코드 출력

    # 파일 선택하기
    def seleteFile(self):
        if 'filename' in self.MainBox.item(self.MainBox.currentRow(), 1).text():  # 기능의 변수가 filename일 때
            if '.' in self.activity_list[self.activity_list.index(self.row_num) + 2]:  # 변수 값에 .이 포함되어있다면 (확장자가 있기 때문에 .을 기준으로함)
                pass
            else:
                name = QFileDialog.getOpenFileName()  # 변수 값에 .이 포함되지 않았다면
                self.activity_list[self.activity_list.index(self.row_num) + 2] = name[0]  # 파일 열기 다이얼로그에서 선택한 경로와 이름으로 프로퍼티 리스트에 값 변경

    # 엑셀, 이미지 미리보기
    def ViewExcelImage(self):
        if 'filename' in self.MainBox.item(self.MainBox.currentRow(), 1).text():  # 기능의 변수가 filename일 때
            value = self.activity_list[self.activity_list.index(self.row_num) + 2]  # 변수 값 불러오기
        else:
            value = ''

        # 엑셀파일
        if '.xlsx' in value:  # 확장자가 엑셀일 때
            try:
                name = value
                wb = xlrd.open_workbook(name)  # 엑셀파일 열기
                ws = wb.sheet_by_index(0)  # 첫번째 시트

                numrow = ws.nrows  # 첫번째 시트의 row
                numcol = ws.ncols  # 첫번째 시트의 col

                self.excelBox.setColumnCount(numcol)  # 엑셀 미리보기 테이블의 col 갯수 설정
                self.excelBox.setRowCount(numrow)  # 엑셀 미리보기 테이블의 row 갯수 설정

                # 엑셀 박스에 데이터 입력
                for row in range(numrow):
                    for col in range(numcol):
                        if ws.row_values(row)[col]:  # 해당 칸에 값이 있을 때
                            self.excelBox.setItem(row, col, QTableWidgetItem(str(ws.row_values(row)[col])))  # 엑셀 미리보기 테이블의 해당칸에 값 입력

            except FileNotFoundError:  # 파일이 없는 에러가 발생할 때
                self.activity_list[self.activity_list.index(self.row_num) + 2] = ''  # 해당 기능의 변수 값 초기화
                self.seleteFile()  # 파일 선택함수 실행
                self.ViewExcelImage()  # 다시 미리보기 함수 실행

        else:
            self.excelBox.setColumnCount(0)  # 엑셀 미리보기 테이블의 col 갯수 0
            self.excelBox.setRowCount(0)  # 엑셀 미리보기 테이블의 row 갯수 0
        # 그림파일
        if '.png' in value or '.jpg' in value or '.jpeg' in value or '.PNG' in value or '.JPG' in value or '.JPEG' in value:  # 확장자가 그림파일일 때
            try:
                pixmap = QPixmap(value).scaled(self.imgLabel.size())  # 이미지라벨의 사이즈로 스케일링
                self.imgLabel.setPixmap(pixmap)  # 이미지 띄우기

            except FileNotFoundError:  # 파일이 없는 에러가 발생할 때
                self.activity_list[self.activity_list.index(self.row_num) + 2] = '' # 해당 기능의 변수 값 초기화
                self.seleteFile()  # 파일 선택함수 실행
                self.ViewExcelImage()  # 다시 미리보기 함수 실행

        else:
            self.imgLabel.setPixmap(QPixmap(None))  # 이미지 없애기

    # 검색기능
    def SearchActivity(self):
        if self.SearchBox.text() == '':  # 검색창의 텍스트가 비었을 때
            seletedlist = self.ActivityBox.selectedItems()  # 선택된 아이템들
            for i in range(len(seletedlist)):
                seletedlist[i].setSelected(False)  # 선택된 아이템들 선택안됨으로 변경
        else:
            self.ActivityBox.expandAll()  # 트리위젯 펼치기
            seletedlist = self.ActivityBox.selectedItems()  # 선택된 아이템들
            for i in range(len(seletedlist)):
                seletedlist[i].setSelected(False)  # 선택된 아이템들 선택안됨으로 변경
            activitylist = self.ActivityBox.findItems(self.SearchBox.text(),Qt.MatchContains|Qt.MatchRecursive)  # 검색한 단어가 포함된 기능찾기
            for i in range(len(activitylist)):
                activitylist[i].setSelected(True)  # 검색된 아이템들 선택됨으로 변경

    def selectPosition(self):
        if 'X, Y' in self.MainBox.item(self.MainBox.currentRow(), 1).text():  # 메인 박스에 X, Y가 있을 경우
            if 'Pos_' in self.activity_list[self.activity_list.index(self.row_num) + 2]:  # X, Y의 값이 Pos_인 경우(초기 값)
                pyautogui.alert(text='마우스 커서의 위치를 반환 = q', title='Mouse Position', button='OK')  # 마우스 커서 위치를 메세지박스로 출력
                while True:
                    if keyboard.is_pressed('q'):  # 키보드의 'q'를 눌렀을 때
                        Activity.pos_X, Activity.pos_Y = pyautogui.position()  # 마우스의 위치를 반환
                        break
                pyautogui.alert(text='x={0}, y={1}'.format(Activity.pos_X, Activity.pos_Y), title='Mouse Position', button='OK')  # 반환된 마우스의 위치를 저장

                self.activity_list[self.activity_list.index('Pos_X')] = str(Activity.pos_X)
                self.activity_list[self.activity_list.index('Pos_Y')] = str(Activity.pos_Y)  # 반환된 마우스의 위치를 엑티비티 리스트에 출력

    # 반복문 기능
    def For_(self, startnum, endnum, roofnum):
        for i in range(int(roofnum)-1):  # 반복횟수에서 -1 한 횟수만큼 반복 (메인박스에서 밑으로 내려가면서 한번 더 동작하기때문)
            for row in range(int(startnum)-1, int(endnum)):  # 시작번호와 끝번호만큼 반복
                self.Action(row)  # 동작 과정함수 실행

    # 사용자함수 생성
    def CreateMyItem(self):
        # 저장
        self.SaveMyItem()  # 649번째 줄 사용자함수 저장기능
        self.plainTextEdit.clear()  # 기존에 있던 사용자함수 정의 칸 초기화
        # 추가
        self.my_item = self.my_item + 1  # 사용자함수의 갯수 +1
        self.my_text.append('')  # 빈칸으로 추가
        self.my_num = self.my_item  # 현재 선택된 사용자함수 번호를 새로 추가된 번호로 변경
        if self.MainBox.rowCount() > 0:  # 메인박스에 아이템이 있을 때
            self.MainBox.insertRow(self.MainBox.rowCount())  # 메인박스 맨 밑줄에 한줄 추가
        else:
            self.MainBox.setRowCount(1)  # 메인박스의 줄 갯수를 1로 변경
        name = 'MyItem' + str(self.my_item)  # 함수 이름을 MyItem + 번호로 지정
        self.MainBox.setItem(self.MainBox.rowCount() - 1, 0, QTableWidgetItem(name))  # 메인박스에 추가되는 기능이름을 name의 값으로 지정
        self.MainBox.setItem(self.MainBox.rowCount() - 1, 1, QTableWidgetItem('_'+name))  # 메인박스에 추가되는 변수를 _ + name의 값으로 지정
        self.activity_list.append(self.MainBox.rowCount() - 1)  # 프로퍼티 리스트에 줄번호를 추가
        self.activity_list.append(name)  # 프로퍼티 리스트에 기능이름을 추가
        self.activity_list.append('_'+name)  # 프로퍼티 리스트에 변수 이름을 추가
        self.MainBox.setCurrentCell(self.MainBox.rowCount() - 1, 0)  # 현재 셀을 추가된 셀로 변경
        if self.varBox.rowCount() > 0:  # 기존에 아이템이 있을 때
            self.SetProperty(self.row_num)  # 기존에 있던 프로퍼티 값을 저장
        self.row_num = self.MainBox.currentRow()  # 줄번호 변수를 현재 줄번호로 변경
        self.ViewExcelImage()  # 기존에 보여지던 엑셀이나 이미지를 감추기위해 실행
        self.ShowProperty(self.MainBox.rowCount() - 1)  # 사용자함수의 프로퍼티를 보여주기

    # 사용자함수 저장
    def SaveMyItem(self):
        if self.my_item >= 0:  # 사용자함수가 하나라도 있을 때
            if 'MyItem' in self.MainBox.item(self.row_num, 0).text():  # 현재 선택된 액티비티가 사용자함수일 때
                self.my_text[self.my_num] = self.plainTextEdit.toPlainText()  # 현재 선택된 사용자함수의 데이터를 저장하기


    # 사용자함수 출력
    def ShowMyItem(self):
        self.plainTextEdit.clear()  # 사용자함수 정의 칸을 초기화
        self.plainTextEdit.setPlainText(self.my_text[self.my_num])  # 현재 선택된 사용자함수의 저장되어있던 데이터를 출력

    # 사용자함수 실행
    def RunMyItem(self, row):
        string = self.MainBox.item(row, 0).text()  # 사용자함수의 이름 가져오기
        self.my_num = int(string[6:])  # 몇번째 함수인지 확인
        my_item_row = self.my_text[self.my_num].split('\n')  # 엔터를 기준으로 나눠서 저장
        for i in range(len(my_item_row)):  # 저장한 배열의 수만큼 반복
            exec(my_item_row[i])  # 텍스트를 코드로 변환하여 실행

app = QApplication(sys.argv)
main_window = MainWindow()
main_window.show()
app.exec_()